import openpyxl as xl

t01 = xl.load_workbook('excel01.xlsx')
sheet01 = t01['Sheet1']
cell01 = sheet01.cell(1, 2)  # 1参表示第几行，2参表示第几列

print('第一行： ', cell01.value)

for row_idx in range(2, sheet01.max_row + 1):  # sheet01.max_row 得到表格的行数
    cell = sheet01.cell
    total_money = cell(row_idx, 2).value * cell(row_idx, 3).value
    cell(row_idx, 4).value = total_money

t01.save('excel02.xlsx')



